
import os
import openpyxl
import natsort
import tempfile
from openpyxl.utils import get_column_letter


def create_excel_file(file_name):
    wb = openpyxl.Workbook()
    path=None
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as file:
        path = file.name
        wb.save(path)
    
    return path

def rename_excel_file(file_path, original_name):
    wb = openpyxl.load_workbook(file_path)
    wb.save(f"{file_path.split("\\\\")[:-1]}original_name")
    return original_name

def create_new_sheet(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    wb.create_sheet(sheet_name)
    wb.save(file_path)
    return f"Sheet '{sheet_name}' created in {file_path}", file_path

def read_excel_file(file_path):
    # Select the active sheet or specify the sheet by name
    wb = openpyxl.load_workbook(file_path)
    
    available_sheets = wb.sheetnames


    # Loop through all rows and cells
    sheets = {}
    
    for sheet in available_sheets:
        sheet = wb[sheet]  # or wb['Sheet1']
        valuable_cells = []
        for row in sheet.iter_rows():
            for cell in row:
                if  cell.value:
                    valuable_cells.append(f"{cell.coordinate}: {cell.value}")

        valuable_cells = natsort.natsorted(valuable_cells)
        sheets[sheet.title] = valuable_cells
    
    result = ""
    for sheet in sheets:
        result += f"SHEET: {sheet}\n"
        result += "\n".join(sheets[sheet]) + "\n"
    print(result)

    return result

def add_formula_to_excel(file_path, sheet_name,column_letter, formula_template, start_row, end_row):
    print(f"Adding formula in column {column_letter} from row {start_row} to {end_row}")
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]  # or wb['Sheet1']
    
    if end_row is None:
        end_row = sheet.max_row
    
    for row in range(start_row, end_row + 1):

        formula = formula_template.format(row=row)
        cell = f"{column_letter}{row}"
        sheet[cell] = formula
    
    wb.save(file_path)
    return f"Formula applied to {column_letter}{start_row}:{column_letter}{end_row} in {file_path}", file_path

def write_to_cell(file_path, sheet_name, column, row, value):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]  # or wb['Sheet1']
    
    cell = f"{column}{row}"
    sheet[cell] = value
    
    wb.save(file_path)
    return f"Value '{value}' written to {cell} in {file_path}", file_path



if __name__ == "__main__" :
    create_excel_file("Hello")